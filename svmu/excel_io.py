from __future__ import annotations

import os
from dataclasses import dataclass
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

DEFAULT_COLUMNS = [
    "id",
    "title",
    "bullets",
    "tags",
    "description",
    "status",
    "output_filename",
    "output_datetime",
]


@dataclass
class IdeaRow:
    idx: int  # DataFrame index
    id: str
    title: str
    bullets: str
    tags: Optional[str]
    description: Optional[str]
    status: str
    output_filename: Optional[str]
    output_datetime: Optional[str]


class ExcelStore:
    def __init__(self, excel_path: str, sheet_name: Optional[str] = None):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        if not os.path.isfile(self.excel_path):
            raise FileNotFoundError(f"Excel not found: {self.excel_path}")

    def read_ready(self, status_ready: str = "Ready") -> List[IdeaRow]:
        df = pd.read_excel(self.excel_path, sheet_name=self.sheet_name)
        # 以降は既存のままでOK（読むだけなら書式は壊れない）
        for col in DEFAULT_COLUMNS:
            if col not in df.columns:
                df[col] = None
        df = df.fillna("")
        ready_df = df[df["status"].astype(str).str.lower() == status_ready.lower()]
        rows: List[IdeaRow] = []
        for idx, r in ready_df.iterrows():
            rows.append(
                IdeaRow(
                    idx=idx,
                    id=str(r["id"]) if str(r["id"]).strip() != "" else str(idx),
                    title=str(r["title"]).strip(),
                    bullets=str(r["bullets"]).strip(),
                    tags=str(r["tags"]).strip() or None,
                    description=str(r["description"]).strip() or None,
                    status=str(r["status"]).strip(),
                    output_filename=str(r["output_filename"]).strip() or None,
                    output_datetime=str(r["output_datetime"]).strip() or None,
                )
            )
        # Keep original df for later writes
        self._df = df
        return rows

    def write_status(
            self,
            row_index: int,
            status_done: str = "Done",
            output_filename: Optional[str] = None,
            output_datetime: Optional[str] = None,
    ) -> bool:
        # openpyxl で既存ブックを開く（書式を保持）
        wb = load_workbook(self.excel_path)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active

        # 1) ヘッダー行（1行目）から列マップを作る
        header_row = 1
        col_map = {}
        max_col = ws.max_column
        for c in range(1, max_col + 1):
            v = ws.cell(row=header_row, column=c).value
            if isinstance(v, str) and v.strip() != "":
                col_map[v.strip()] = c

        # 2) 必須列が無ければ末尾に追加
        changed_header = False
        for col_name in DEFAULT_COLUMNS:
            if col_name not in col_map:
                max_col += 1
                ws.cell(row=header_row, column=max_col).value = col_name
                col_map[col_name] = max_col
                changed_header = True
        # ヘッダー追加のみでは書式はほぼ影響なし（列幅を変えないため）

        # 3) 書き込み先のExcel行番号を算出（ヘッダー1行分オフセット）
        target_row = row_index + 2

        # 4) 値のみ更新（書式は既存セルのものを維持）
        if "status" in col_map:
            ws.cell(row=target_row, column=col_map["status"]).value = status_done
        if output_filename is not None and "output_filename" in col_map:
            ws.cell(row=target_row, column=col_map["output_filename"]).value = output_filename
        if output_datetime is not None and "output_datetime" in col_map:
            # 文字列として書き込む（書式崩れ防止）
            ws.cell(row=target_row, column=col_map["output_datetime"]).value = output_datetime
            # 日時を数値で持ちたい場合は以下のように datetime を設定し、number_format を指定
            # from datetime import datetime
            # cell = ws.cell(row=target_row, column=col_map["output_datetime"])
            # cell.value = datetime.now()
            # cell.number_format = "yyyy/mm/dd hh:mm:ss"

        # 5) 保存
        try:
            wb.save(self.excel_path)
            return True
        except Exception as e:
            return False
